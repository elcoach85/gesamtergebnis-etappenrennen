import re
import logging
from dataclasses import dataclass
from pathlib import Path

import datetime
import fuzzpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

get_cell_rowidx = lambda cell: cell.row
get_cell_value = lambda cell: cell.value


@dataclass
class RowStartIdx:
    row_start_idx: int
    name: str
    level: int
    description: str
    row_end_idx: int = None

    def map_to_data_idxs(self, data_idxs):
        import numpy as np
        
        tmp = np.digitize(data_idxs, [self.row_start_idx, self.row_end_idx])
        data_idxs = np.array(data_idxs)
        return data_idxs[tmp == 1]


def read_row(ws, row_idx):
    tmp = [
        value.strftime("%H:%M:%S") if isinstance(value, datetime.time) else value
        for row in ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True)
        for value in row
    ]
    return tmp


def read_rows(ws, row_list):
    return [read_row(ws, row_idx) for row_idx in row_list]


def row_start_idx_from_list_of_cells(ls, level, description):
    return [RowStartIdx(cell.row, cell.value, level, description) for cell in ls]


def get_raceresult_from_ws(ws, search_rating_category):

    first_coloumn_area = fuzzpyxl.CellArea(max_col=2)
    categorie_idxs = fuzzpyxl.find_values_in_area(
        worksheet=ws,
        search_cell=lambda cell: cell.value in search_rating_category,
        search_area=first_coloumn_area,
    )

    if len(categorie_idxs) == 0:
        raise ValueError(
            f"Can not find classname of tyhe provided {search_rating_category} in worksheet {ws.title}"
        )

    dnf_idxs = fuzzpyxl.find_values_in_area(
        worksheet=ws, search_cell="DNF", search_area=first_coloumn_area
    )
    dns_idxs = fuzzpyxl.find_values_in_area(
        worksheet=ws, search_cell="DNS", search_area=first_coloumn_area
    )

    # Combine and sort
    row_areas1 = row_start_idx_from_list_of_cells(
        categorie_idxs, 1, "wertungskategorie"
    )
    row_areas2 = row_start_idx_from_list_of_cells(dnf_idxs, 2, "kein_valides_ergebnis")
    row_areas3 = row_start_idx_from_list_of_cells(dns_idxs, 2, "kein_valides_ergebnis")

    re_uci_number = re.compile(r"^\d{1,3}$")
    uci_numbers = fuzzpyxl.find_values_in_area(
        worksheet=ws,
        search_cell=lambda cell: re_uci_number.match(str(cell.value)),
        search_area=fuzzpyxl.CellArea(min_col=2, max_col=2),
    )

    data_rowidxs = list(map(get_cell_rowidx, uci_numbers))


    pos_name_cell = fuzzpyxl.find_first_value_in_area(
        ws, "Pos.", fuzzpyxl.CellArea(min_col=1, max_col=1, max_row=10)
    )
    column_names = read_row(ws, pos_name_cell.row)

    row_areas = [*row_areas1, *row_areas2, *row_areas3]
    row_areas = sorted(row_areas, key=lambda el: el.row_start_idx, reverse=True)
    unique_levels = sorted(list(set([el.level for el in row_areas])))

    ws_max_row = max(data_rowidxs) + 1 if ws.max_row is None else ws.max_row + 1

    prev_idx_per_level = {level: ws_max_row for level in unique_levels}

    for el in row_areas:
        logging.debug(prev_idx_per_level)
        el.row_end_idx = prev_idx_per_level[el.level]
        for level in prev_idx_per_level.keys():
            if level >= el.level:
                prev_idx_per_level[level] = el.row_start_idx

    row_areas = sorted(row_areas, key=lambda el: el.row_start_idx)

    logging.debug(f"wertungskategorie: {row_areas}")

    dfs = []

    for row_area in filter(lambda el: el.level == 1, row_areas):
        row_idxs = row_area.map_to_data_idxs(data_rowidxs)
        data = read_rows(ws, row_idxs)
        df = pd.DataFrame(data, index=row_idxs, columns=column_names)
        df[row_area.description] = str(row_area.name)
        for row_area in filter(lambda el: el.level > 1, row_areas):
            row_idxs = row_area.map_to_data_idxs(data_rowidxs)
            row_idxs = list(set(df.index).intersection(row_idxs))
            if not row_area.description in df.columns:
                df[row_area.description] = ""
            df.loc[row_idxs, row_area.description] = row_area.name
        dfs.append(df)

    df = pd.concat(dfs)
    # assert pd.notna(df.iloc[0]['Team']),
    return df


def _process_ws(ws: Worksheet, search_rating_category):
    logging.info(f"processing {ws.title}")
    df = get_raceresult_from_ws(ws, search_rating_category)
    return df


def extract_race_result_from_ws(wb_path, search_rating_category, not_start_handicap):
    wb_path = Path(wb_path)
    wb = load_workbook(filename=str(wb_path), read_only=True)
    df = pd.concat([_process_ws(ws, search_rating_category) for ws in wb.worksheets])

    df["time_in_s"] = pd.to_timedelta(df["Gesamtzeit"]).apply(
        lambda el: el.total_seconds()
    )

    print("passed")
    df.loc[df["Gesamtzeit"].isna(), "time_in_s"] = not_start_handicap
    return df
