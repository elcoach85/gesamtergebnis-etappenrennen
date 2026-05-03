import re
from pathlib import Path


from openpyxl import load_workbook
from openpyxl.cell.read_only import EmptyCell
import fuzzpyxl
import pandas as pd


def get_row_as_list(ws, race_name_cell, row_padding=1, col_padding=1):
    res = []
    for row in ws.iter_rows(
        min_row=race_name_cell.row + row_padding,
        max_row=race_name_cell.row + row_padding,
        min_col=race_name_cell.column + col_padding,
    ):

        for cell in row:
            if isinstance(cell, EmptyCell):
                return res
            res.append(cell.value)

        return res


def _process_ws(ws, race_locations):

    race_name = ws.cell(1, 1).value

    marking = re.compile(r"x")

    dfs = []

    for location_name in race_locations:
        race_name_cell = fuzzpyxl.find_first_value_in_area(
            ws, location_name, fuzzpyxl.CellArea(max_row=5)
        )
        wertungsnamen = get_row_as_list(ws, race_name_cell, 1)
        wertungseinheiten = get_row_as_list(ws, race_name_cell, 2)
        wertungsgutschriften = get_row_as_list(ws, race_name_cell, 3)
        onlydigits = re.compile(r"\d+")
        wertungsgutschriften = [
            list(onlydigits.findall(wertungsgutschrift))
            for wertungsgutschrift in wertungsgutschriften
        ]
        # bekomme die runden

        rundezahl = 0
        rundenderwertungen = list()

        for row in ws.iter_rows(
            min_col=race_name_cell.column,
            max_col=race_name_cell.column + len(wertungsnamen) + 1,
            min_row=race_name_cell.row + 4,
        ):
            for rel_idx, cell in enumerate(row):
                if rel_idx == 0:
                    rundezahl = cell.value

                if bool(marking.match(str(cell.value))) and not rel_idx == 0:
                    wertungstyp = wertungsnamen[rel_idx - 1]
                    wertungsgutschrift = wertungsgutschriften[rel_idx - 1]
                    wertungseinheit = wertungseinheiten[rel_idx - 1]
                    for pos, gutshrift in enumerate(wertungsgutschrift, start=1):
                        rundenderwertungen.append(
                            {
                                "renn_ort": location_name,
                                "wertungskategorie": race_name,
                                "wertungstyp": wertungstyp,
                                "rundenanzahl": rundezahl,
                                "position": pos,
                                "gutschrift": gutshrift,
                                "einheit": wertungseinheit,
                            }
                        )

        dfs.append(pd.DataFrame(rundenderwertungen))

    df = pd.concat(dfs)
    return df


def read_point_score_file(wb_path, race_locations, punktexcel2wertungskategorie):
    wb_path = Path(wb_path)
    wb = load_workbook(filename=str(wb_path), read_only=True)
    df = pd.concat([_process_ws(ws, race_locations) for ws in wb.worksheets])
    df["wertungskategorie"] = df["wertungskategorie"].map(punktexcel2wertungskategorie)
    return df


def get_location_racename(ws):
    location, race_name = ws["A1"].value.split("_")
    return location, race_name


def read_template_score_file(wb_path):
    wb_path = Path(wb_path)
    wb = load_workbook(filename=str(wb_path), read_only=True)
    title2locrcen = {ws.title: get_location_racename(ws) for ws in wb.worksheets}

    dfs = []
    for ws_title, (location, race_name) in title2locrcen.items():
        df = pd.read_excel(wb_path, sheet_name=ws_title, skiprows=[0, 1])
        df["renn_ort"] = location
        df["wertungktegorie"] = race_name
        df = df.dropna(subset=["startnummer", "transpondernummer"], how="all")
        dfs.append(df)

    df = pd.concat(dfs)
    return df
