import argparse
from openpyxl.styles import Font
import pandas as pd
import racedaystools
import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter


def set_margins(filename, margins_dct):
    wb = load_workbook(filename)
    for ws in wb.worksheets:
        ws.page_margins = PageMargins(**margins_dct)

    wb.save(filename)


def adjust_coloumn_widths(writer, sheet_name, df, extra_padding=4):
    for column in df:
        column_length = max(df[column].astype(str).map(len).max(), len(column))
        column_length += extra_padding
        print(column, column_length)
        col_idx = df.columns.get_loc(column)
        writer.sheets[sheet_name].column_dimensions[
            get_column_letter(col_idx + 2)
        ].width = column_length


def make_team_results_from_gesamt(
    db_engine, file_name, top_riders_per_category, with_rounds
):
    with db_engine.connect() as conn:
        time_colname = "total_time_runden_boni" if with_rounds else "total_time_boni"

        df = pd.read_sql(
            f"SELECT g.wertungskategorie, t.verein_team , g.{time_colname} AS race_time "
            "FROM gesamt_zeit_wertung g "
            "INNER JOIN teilnehmer t ON g.startnummer = t.startnummer",
            conn,
        )
        sheetname2df = {}
        with pd.ExcelWriter(file_name, mode="w") as writer:
            for category, n_riders in top_riders_per_category.items():
                df_tmp = df[df["wertungskategorie"] == category]
                df_tmp = df_tmp.sort_values(["race_time"])
                df_top_riders = (
                    df_tmp.groupby(["verein_team"])
                    .head(n_riders)
                    .reset_index(drop=True)
                )

                s_result = df_top_riders.groupby("verein_team")["race_time"].sum()

                df_result = s_result.to_frame()

                df_result["Teamname"] = df_result.index.to_list()
                df_result["Gesamtzeit"] = df_result["race_time"].apply(
                    lambda el: str(datetime.timedelta(seconds=el))
                )

                df_result = df_result.sort_values(["race_time"], ignore_index=True)
                df_result = df_result.drop(columns=["race_time"])

                ws_name = category
                df_result.index = list(range(1, df_result.shape[0] + 1))
                df_result.to_excel(
                    writer,
                    sheet_name=ws_name,
                    index=True,
                    index_label="Pos.",
                    startrow=2,
                )
                ws = writer.sheets[ws_name]
                ws["A1"] = ws_name
                ws["A1"].font = Font(bold=True, sz=14)

                adjust_coloumn_widths(writer=writer, sheet_name=ws_name, df=df_result)
                sheetname2df[ws_name] = df_result

            df = pd.concat(sheetname2df.values())
            for sheet_name in sheetname2df:
                adjust_coloumn_widths(writer=writer, sheet_name=sheet_name, df=df)


def make_team_results_from_daily(
    db_engine, file_name, top_riders_per_category, with_rounds
):
    with db_engine.connect() as conn:
        time_colname = (
            "time_in_s_mit_runden_boni" if with_rounds else "time_in_s_mit_boni"
        )

        df = pd.read_sql(
            f"SELECT e.startnummer, t.verein_team ,e.renn_ort,e.wertungskategorie , e.{time_colname} AS race_time "
            "FROM ergebnis_mit_runden_boni e "
            "INNER JOIN teilnehmer t ON e.startnummer = t.startnummer",
            conn,
        )

        sheetname2df = {}
        with pd.ExcelWriter(file_name, mode="w") as writer:
            for category, n_riders in top_riders_per_category.items():
                df_tmp = df[df["wertungskategorie"] == category]
                df_tmp = df_tmp.sort_values(["race_time"])
                df_top_riders = (
                    df_tmp.groupby(["verein_team", "renn_ort"])
                    .head(n_riders)
                    .reset_index(drop=True)
                )

                s_result = df_top_riders.groupby("verein_team")["race_time"].sum()

                df_result = s_result.to_frame()

                df_result["Teamname"] = df_result.index.to_list()
                df_result["Gesamtzeit"] = df_result["race_time"].apply(
                    lambda el: str(datetime.timedelta(seconds=el))
                )

                df_result = df_result.sort_values(["race_time"], ignore_index=True)
                df_result = df_result.drop(columns=["race_time"])

                ws_name = category
                df_result.index = list(range(1, df_result.shape[0] + 1))
                df_result.to_excel(
                    writer,
                    sheet_name=ws_name,
                    index=True,
                    index_label="Pos.",
                    startrow=2,
                )
                ws = writer.sheets[ws_name]
                ws["A1"] = ws_name
                ws["A1"].font = Font(bold=True, sz=14)
                sheetname2df[ws_name] = df_result

            df = pd.concat(sheetname2df.values())
            for sheet_name in sheetname2df:
                adjust_coloumn_widths(writer=writer, sheet_name=sheet_name, df=df)


def make_team_results(
    calculation_method, db_engine, file_name, top_riders_per_category, with_rounds
):
    if calculation_method == "tages":
        make_team_results_from_daily(
            db_engine=db_engine,
            file_name=file_name,
            top_riders_per_category=top_riders_per_category,
            with_rounds=with_rounds,
        )
    elif calculation_method == "gesamt":
        make_team_results_from_gesamt(
            db_engine=db_engine,
            file_name=file_name,
            top_riders_per_category=top_riders_per_category,
            with_rounds=with_rounds,
        )


def make_points_resultsfile(db_engine, file_name):
    # query the database
    with db_engine.connect() as conn:
        df = pd.read_sql("SELECT * FROM gesamt_punkte_wertung", conn)

    with pd.ExcelWriter(file_name, mode="w") as writer:
        for grp_name, df_grp in df.groupby(["wertungskategorie", "wertungstyp"]):
            wertungskategorie, wertungstyp = grp_name
            ws_name = f"{wertungskategorie}_{wertungstyp}"
            coloumns = df_grp.columns.to_list()
            coloumns.remove("wertungstyp")
            coloumns.remove("wertungskategorie")
            df_grp.index = list(range(1, df_grp.shape[0] + 1))
            df_grp.to_excel(
                writer,
                sheet_name=ws_name,
                index=True,
                index_label="Pos.",
                startrow=2,
                columns=coloumns,
            )
            ws = writer.sheets[ws_name]
            ws["A1"] = ws_name
            ws["A1"].font = Font(bold=True, sz=14)

            adjust_coloumn_widths(writer=writer, sheet_name=ws_name, df=df[coloumns])


def make_individual_resultsfile(db_engine, file_name, with_rounds):
    with db_engine.connect() as conn:
        total_time_colname = (
            "total_time_runden_boni" if with_rounds else "total_time_boni"
        )
        df = pd.read_sql(
            f"SELECT wertungskategorie,startnummer,name,vorname,{total_time_colname} AS total_time "
            "FROM gesamt_zeit_wertung",
            conn,
        )

        df["Gesamtzeit"] = df["total_time"].apply(
            lambda el: str(datetime.timedelta(seconds=el))
        )
        df = df.sort_values(["total_time"], ascending=True)
        df = df.drop(columns=["total_time"])

    with pd.ExcelWriter(file_name, mode="w") as writer:
        for grp_name, df_grp in df.groupby("wertungskategorie"):
            ws_name = grp_name
            coloumns = df_grp.columns.to_list()
            coloumns.remove("wertungskategorie")

            df_grp.index = list(range(1, df_grp.shape[0] + 1))
            df_grp.to_excel(
                writer,
                sheet_name=ws_name,
                index=True,
                index_label="Pos.",
                startrow=2,
                columns=coloumns,
            )
            ws = writer.sheets[ws_name]
            ws["A1"] = ws_name
            ws["A1"].font = Font(bold=True, sz=14)

            adjust_coloumn_widths(writer=writer, sheet_name=ws_name, df=df[coloumns])


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Try to import all files into the database"
    )
    parser.add_argument(
        "-config_path", type=str, default="config.yaml", help="path to the config file"
    )
    args = parser.parse_args()

    cfg = racedaystools.read_setup_yaml(args.config_path)
    # Create database engine
    db_engine = racedaystools.get_sql_engine(cfg)

    make_points_resultsfile(db_engine, cfg["Outputfiles"]["punkte"])

    make_individual_resultsfile(
        db_engine,
        cfg["Outputfiles"]["gesamt"],
        cfg["Outputfiles"]["mit_berechnung_rundenrueckstand"],
    )
    make_team_results(
        cfg["Outputfiles"]["berechnungsmethode_team"],
        db_engine,
        cfg["Outputfiles"]["team"],
        cfg["Outputfiles"]["numberbestriderperrating_category"],
        cfg["Outputfiles"]["mit_berechnung_rundenrueckstand"],
    )

    # make set_margins
    set_margins(cfg["Outputfiles"]["team"], cfg["ExcelFormat"]["margins"])
    set_margins(cfg["Outputfiles"]["punkte"], cfg["ExcelFormat"]["margins"])
    set_margins(cfg["Outputfiles"]["gesamt"], cfg["ExcelFormat"]["margins"])
